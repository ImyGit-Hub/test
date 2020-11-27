import tweepy
import openpyxl

#Excel読み込み
wb=openpyxl.load_workbook("TweetByPython.xlsx")
Tweetsheet=wb["TweetList"]
TweetsheetLastRow=Tweetsheet.max_row

# 自動ツイートに必要となる各種キーを代入する
#CK="Consumer Key(API key)"
#CS="Consumer Secret(API secret)"
#AT="Access Token"
#AS="Access Token Secret"

#https://developer.twitter.com/en/portal/dashboard
#https://qiita.com/butsuli_shine/items/78fd5ee6fdb4a0581652

CK=""
CS=""
AT=""
AS=""

# Twitterオブジェクトの生成
auth = tweepy.OAuthHandler(CK, CS)
auth.set_access_token(AT, AS)
api = tweepy.API(auth)

for row in range(2,TweetsheetLastRow+1):

    if Tweetsheet.cell(row,2).value == None:
        #画像なしツイート
        api.update_status(Tweetsheet.cell(row,1).value)
        print("画像なし:"+ Tweetsheet.cell(row,1).value)
    else:
        #画像付きツイート
        api.update_with_media(status = Tweetsheet.cell(row,1).value, filename = Tweetsheet.cell(row,2).value)
        print("画像付き:"+ Tweetsheet.cell(row,1).value)

print("完了")


